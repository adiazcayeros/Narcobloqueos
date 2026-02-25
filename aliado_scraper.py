"""
Aliado Neuralgic Alerts Scraper
Fetches full details for all alerts and saves to Excel.
"""

import requests
import time
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
API_URL = "https://alertas.aliado.alephri.com/api/graphql"
BEARER_TOKEN = "e2e29d3164218e91ea40dd6c63808b79"
OUTPUT_FILE = "aliado_alerts_feb22_2026.xlsx"

HEADERS = {
    "Authorization": f"Bearer {BEARER_TOKEN}",
    "Content-Type": "application/json",
}

# All 138 alert IDs from GET_NEURALGIC_ALERTS_MARKERS
ALERT_IDS = [
    "878fb9bc-02a3-4abc-8b68-378291a05baf","651284db-5c73-4686-919c-0b26484329b5",
    "9603a240-5bf1-4b88-a9b2-3e2758098ad8","e72c2b5e-776f-45b5-8747-7ff66144da14",
    "0a9f3649-1c9d-440b-a00b-8f3971f90a55","56a7edcb-6ea3-4415-9ba7-671443f8a6f5",
    "6664aa95-33de-458a-bbcb-beb8a099f730","27f22799-cc63-4917-a5fa-886d75271096",
    "98ea7abe-ff66-4b5f-ac36-10a49a8abdba","62ec7d64-3a57-4531-ac1e-1268ba7bfc14",
    "971ac6b0-5c84-47ea-9b98-dbbb2623da90","3a8d2e35-be84-4fa0-b6fc-801ca29cb643",
    "2c85791f-0f13-4bb8-9824-186828b03b8b","2d2a4cba-ef1e-4756-b736-64b6f6cda1cf",
    "2a193565-c3d7-4b74-b0e4-e8009ee1aaf9","3a7ebe67-fbce-4583-bb26-af819aab68b3",
    "23aaf340-e5c1-4092-9774-d915f48311f7","2dce7041-47e5-4b14-922b-3dbdf1f26d91",
    "db323d1f-3c89-4008-b4e2-f7e76ce82a47","8e1f2776-6c69-4f5f-8b16-bb522f00f839",
    "8a44cf15-d75b-4fb7-9e12-4a178b7f5276","ae7e997a-ba1c-4198-87b4-f82a727c9f7c",
    "537f2e50-2e0b-4630-9835-230d9494c391","39e3f8a2-ddc7-4183-a7ae-048f10637c00",
    "87e1a8a0-ae2e-424f-94a6-08328562b166","c860b943-6e48-4e72-abf4-24a1713841be",
    "25eaf4dd-9fe2-4865-a859-9cce719eb546","e372700f-9f90-466f-9273-b6dbc3317d54",
    "12d308c6-09d1-437c-a4d1-0150910b7544","55d0a3e6-acf4-49ca-af47-f26388ab4a66",
    "03c2c46d-f413-40aa-bdf0-305a42e37910","e828803e-8028-46f2-bd76-a8e5c6cd484f",
    "e81a1337-dea7-4980-b3d6-e7c6f85d58c5","71b203df-e1ad-4f1a-9b55-bd7c8ee4f3bb",
    "b89f2298-b7b9-4c5b-9534-32ea44107cba","993c9f7d-508b-442b-b23d-9a9f3a9bffee",
    "15000370-7e85-4d3b-91c6-64df9c2af1e9","856083eb-4c4c-462a-af3e-04a2df328035",
    "0b6afb7c-aab3-48d4-b69e-9befca3be01f","a71c6106-8dca-49b0-a2cc-c71e0f992de0",
    "b528e5da-c475-4ebf-bb79-700ddf89b4d7","2608ad08-f608-4675-be0f-daa7f22a9471",
    "9e83e554-357d-443d-b61d-88cba6bdc6b5","ac89fa86-5542-4ce8-b64c-c995bd1341a3",
    "fea320a2-55e2-4d4c-b3eb-7eff77447ee6","fc4b5b73-6695-4759-bcd7-047230b4df53",
    "281414c4-9f0a-4e1a-8e1d-de23f6981897","2806faa8-0635-4f88-a732-4db0d284099f",
    "aae450b5-ab75-41b6-a4af-b670bd090063","b975bd6b-caf8-40c1-b8eb-e1e8bf1a2079",
    "e131277c-9daf-436e-a54c-70b2926b460d","8679c380-033a-47a7-830e-f6e3a7976ee8",
    "7885f3d2-8272-4bf5-876a-8b55c7597004","37cf9694-f22c-4455-84a1-0b535611f852",
    "eb13aeb7-fbf5-442e-88c8-a28b44bece8d","240d9e90-1732-4649-a79c-1546ac4f3b3a",
    "39385b35-5c54-4b90-ab14-2e5f9f398bd9","5233e14a-7db3-48cd-8817-6882e8ce3bcb",
    "5011eb17-d54f-4106-ad8b-fae2e9ec9264","226cb998-d4e4-4587-bb00-0dab40c321c4",
    "5fbf647c-6dfb-4f47-80e1-4ae34fc4e4e0","17b080a8-5f7b-4111-9b85-90d0a2a0baf6",
    "63feac85-4227-417d-8268-43b8d00765e5","51363980-9dbb-4185-8793-63abc0cafe3a",
    "41dbbb74-bb7e-4c02-a10d-2730fdcd2532","e5ac8ad7-2c11-43b1-845c-ee8e568e5a64",
    "a9825911-cc43-4684-bc65-617cfab5b5d9","c0b7516f-2308-4079-a4eb-cdaaaa64503b",
    "9c3f522c-98d1-4870-aeed-6a648e41d841","4b29980d-26a9-42c2-9602-6d937ad7989a",
    "67238415-a8f6-4594-84b7-aefb455720fe","7504a5f6-6ffa-4f14-a40d-fc65a1e44aa1",
    "0a0733ee-18de-42dd-9faf-52c64d1e333a","da8b2f4d-3b58-4c01-bde5-fc77b36a48ab",
    "83ab4065-05bc-4273-8d71-1277f5a41d95","4af280ab-e87a-4ef8-a978-e72b4bcce358",
    "02e6fa87-c9a1-4a16-aa9e-5bc76b630c69","6ef7577e-9316-405d-8e5a-03a3e0ffd144",
    "49fbc1f0-66cf-4b22-85f5-5a307195540b","0bd66ce6-a169-4865-b21e-a99f116169bb",
    "8fd229dc-60d1-4e06-a6a4-60d5ecbfb293","6c313619-2e70-42b0-a83a-d999696d5b92",
    "3a642ca4-18a3-448f-9162-832e0b5848f2","b7ecabf3-6e48-4518-8ebd-412e4d0e2f67",
    "2e54af23-55d6-4cb5-9da6-04e18b05152c","16b67632-aa34-441f-83e3-651690352676",
    "702cd9b0-1d93-4eee-8d41-c1e0a4d7eb38","54ae063b-977d-4a38-87ce-082bb6c1a847",
    "e22fbc63-30cd-478b-8b21-10e85ed803cb","275ccc1a-2f8a-4d1e-880a-b8286fe5cbdd",
    "123bc9fb-d5c9-4657-8bbe-1ee8c685de91","30ecb03c-60ca-4fce-9037-ad81e6723b77",
    "4156e311-b47b-4f35-a89b-2196a4fdd391","949ef07b-84d5-43d0-b56e-f18ba12ee598",
    "bfdf9162-9ee8-4a17-b3f1-143b47237e5f","10fdafbe-2979-4737-95dc-ff12386e3613",
    "338cc93b-caa4-4103-9933-288879209fbe","ea410eb2-a376-4297-b6ea-04868695a628",
    "2b3d9ec2-7415-4aea-9ad5-ab7db1e5fa37","5eb806b5-6b15-49af-9212-50dda25ab00a",
    "eb6bf4bd-dc9e-41f9-ae8b-e5fd22ef7bd0","1d9f80aa-788b-4435-ab97-19b89ff5ecb7",
    "558bb89f-7023-4766-b08a-17846102b4ac","4784e710-a090-46b0-935e-b372fe97277f",
    "0c74fe00-8868-45b9-a567-5867026eb2d1","bc470d34-c8b8-4a3e-bb48-9247c59bd720",
    "51b0c381-c9b0-4de4-8a44-04e855808b54","81a78ba1-135f-4ef6-968d-4141f92745f9",
    "141cbbd4-fc35-49d8-83da-a8688a30309b","76178725-9e85-4722-a4f6-dfb097a2f37e",
    "774f6ba6-9660-4c61-b395-8c197e01dc8b","f2a30387-305c-4a1f-b0ed-9c1af9c6b088",
    "2a88e703-fd7e-4a9f-8f8a-2b3c948305bb","a56433c3-3c41-46f0-8090-8bb57c3da13e",
    "23c3e1ab-2e89-4dbc-9fe4-1c66f06ddb8e","f689d298-59fb-44ef-97f4-4472083e9afe",
    "016a6f18-d952-42ea-b3b3-953d16a135b2","5e0045d8-8397-4d19-9d81-8abe2047726f",
    "9615ec39-7d36-467a-8558-ef5eee1e9a26","c77611e7-682f-4e33-a4eb-293a160f8575",
    "78b185fe-2568-4693-a1d7-e546f65a7bbf","7cbfb908-a84d-4b1f-a671-ec23a7a4f9f4",
    "5d47c0e0-a03b-4890-8551-a2cd9f4f2bb5","756af627-1c29-400d-9e92-8a0f99a2bee1",
    "fe4eeaf7-8f94-4ba8-bad6-dffedc12e922","91015fe9-5c25-4a33-a7f5-67e37e69748d",
    "6527268e-569d-40c0-96b3-d6933df1af61","405eb2ab-617e-488e-b5e1-b2c1cf4a8317",
    "2b63cd0e-cd5e-4748-b277-31c6cc418e40","68f612dc-5174-40d8-af70-772b4bd62f9a",
    "3a3f6b65-57ef-4896-b217-9e1d1d1df4a7","b7127684-a7b3-4d2d-8868-73da1e676b88",
    "bfc73636-3c67-4d6a-87d7-08c8d63fcb88","59deeac0-30de-41f3-8c7b-304818e832a7",
    "0961e41e-50c6-49f7-a146-c79e29a49afa","5a8c00fb-2a41-4041-a907-70ade5ad5a38",
    "975675eb-8bf0-4b54-a478-ad79cdcb6720","9fd6da9b-717a-4476-bc65-19c31f1d5b7b",
]

DETAIL_QUERY = """
fragment NeuralgicAlertCoreFields on NeuralgicAlert {
  id title description latlon { lat lon __typename } __typename
}
fragment NeuralgicAlertDetailFields on NeuralgicAlert {
  ...NeuralgicAlertCoreFields
  startDatetime endDatetime distanceInMeters status sourceAsText
  alertType { id name __typename }
  __typename
}
query GET_NEURALGIC_ALERT($id: ID!) {
  neuralgicAlert(id: $id) {
    ...NeuralgicAlertDetailFields
    hasExpired
    __typename
  }
}
"""

def fetch_alert(alert_id):
    payload = {
        "operationName": "GET_NEURALGIC_ALERT",
        "variables": {"id": alert_id},
        "query": DETAIL_QUERY
    }
    try:
        r = requests.post(API_URL, headers=HEADERS, json=payload, timeout=15)
        r.raise_for_status()
        data = r.json()
        return data.get("data", {}).get("neuralgicAlert")
    except Exception as e:
        print(f"  ERROR on {alert_id}: {e}")
        return None

def build_excel(alerts):
    wb = Workbook()
    ws = wb.active
    ws.title = "Alerts"

    HDR_FILL = PatternFill("solid", start_color="1F3864")
    SAFETY_FILL = PatternFill("solid", start_color="FCE5CD")
    VIAL_FILL = PatternFill("solid", start_color="FFF2CC")
    GENERAL_FILL = PatternFill("solid", start_color="D9EAD3")
    ALT_FILL = PatternFill("solid", start_color="EBF0FA")
    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    headers = ["#", "Title", "Alert Type", "Status", "Start", "End",
               "Latitude", "Longitude", "Source", "Description", "ID"]
    widths =  [5,   35,      20,           12,       20,      20,
               12,          12,             25,        80,             38]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    type_fills = {"SAFETY": SAFETY_FILL, "VIAL": VIAL_FILL, "GENERAL": GENERAL_FILL}

    for i, alert in enumerate(alerts, 1):
        if not alert:
            continue

        def fmt_dt(s):
            if not s:
                return ""
            try:
                return datetime.fromisoformat(s.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M")
            except:
                return s

        alert_type_name = (alert.get("alertType") or {}).get("name", "")
        status = alert.get("status", "")
        lat = (alert.get("latlon") or {}).get("lat", "")
        lon = (alert.get("latlon") or {}).get("lon", "")

        row = [
            i,
            alert.get("title", ""),
            alert_type_name,
            status,
            fmt_dt(alert.get("startDatetime")),
            fmt_dt(alert.get("endDatetime")),
            lat, lon,
            alert.get("sourceAsText", ""),
            alert.get("description", ""),
            alert.get("id", ""),
        ]

        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        row_fill = type_fills.get(status, ALT_FILL if i % 2 == 0 else None)
        if row_fill:
            for col in range(1, len(headers)+1):
                ws.cell(row=i+1, column=col).fill = row_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(alerts)+1}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Aliado — Neuralgic Alerts Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3864")
    ws2["A3"] = "Date Range:"
    ws2["B3"] = "Feb 22–23, 2026"
    ws2["A4"] = "Total Alerts:"
    ws2["B4"] = len([a for a in alerts if a])
    ws2["A5"] = "Source:"
    ws2["B5"] = "Aliado (alertas.aliado.alephri.com)"

    type_counts = {}
    status_counts = {}
    for a in alerts:
        if not a:
            continue
        t = (a.get("alertType") or {}).get("name", "Unknown")
        s = a.get("status", "Unknown")
        type_counts[t] = type_counts.get(t, 0) + 1
        status_counts[s] = status_counts.get(s, 0) + 1

    ws2["A7"] = "By Alert Type"
    ws2["A7"].font = Font(name="Arial", bold=True, size=11, color="1F3864")
    ws2["A8"] = "Type"
    ws2["B8"] = "Count"
    for cell in ["A8", "B8"]:
        ws2[cell].fill = HDR_FILL
        ws2[cell].font = Font(name="Arial", bold=True, color="FFFFFF")
    for r, (t, c) in enumerate(sorted(type_counts.items(), key=lambda x: -x[1]), 9):
        ws2.cell(row=r, column=1, value=t)
        ws2.cell(row=r, column=2, value=c)

    ws2["A15"] = "By Status"
    ws2["A15"].font = Font(name="Arial", bold=True, size=11, color="1F3864")
    ws2["A16"] = "Status"
    ws2["B16"] = "Count"
    for cell in ["A16", "B16"]:
        ws2[cell].fill = HDR_FILL
        ws2[cell].font = Font(name="Arial", bold=True, color="FFFFFF")
    for r, (s, c) in enumerate(sorted(status_counts.items(), key=lambda x: -x[1]), 17):
        ws2.cell(row=r, column=1, value=s)
        ws2.cell(row=r, column=2, value=c)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15

    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")

# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"Fetching details for {len(ALERT_IDS)} alerts...\n")
    alerts = []
    for i, aid in enumerate(ALERT_IDS, 1):
        print(f"[{i}/{len(ALERT_IDS)}] {aid}")
        alert = fetch_alert(aid)
        alerts.append(alert)
        time.sleep(0.2)  # polite delay

    success = len([a for a in alerts if a])
    print(f"\nFetched {success}/{len(ALERT_IDS)} alerts successfully.")
    print("Building Excel file...")
    build_excel(alerts)
    print("Done! Open aliado_alerts_feb22_2026.xlsx")
